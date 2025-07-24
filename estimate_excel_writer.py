from openpyxl import load_workbook
import os
from datetime import datetime
import re

def get_customer_address_from_session():
    """セッション状態から顧客住所を取得する（細分化対応）"""
    try:
        import streamlit as st
        # 新しい形式の住所を優先
        郵便番号 = st.session_state.get("選択された郵便番号", "")
        住所1 = st.session_state.get("選択された住所1", "")
        住所2 = st.session_state.get("選択された住所2", "")
        
        return 郵便番号, 住所1, 住所2
    except:
        return "", "", ""

def parse_address(address):
    """住所を郵便番号、住所1（番地まで）、住所2（建物名など）に分割（修正版）"""
    if not address:
        return "", "", ""
    
    # 郵便番号パターン（〒123-4567 または 123-4567）
    postal_pattern = r'〒?(\d{3}-\d{4})'
    postal_match = re.search(postal_pattern, address)
    
    郵便番号 = ""
    残り住所 = address
    
    if postal_match:
        郵便番号 = postal_match.group(1)
        残り住所 = address.replace(postal_match.group(0), "").strip()
    
    # 住所1（番地まで）と住所2（建物名など）を分割
    住所1 = ""
    住所2 = ""
    
    # **重要：建物名が明確にある場合のみ分割する**
    # 建物名パターン（明確な建物名キーワードがある場合のみ）
    building_patterns = [
        # 明確な建物名キーワード + その他
        r'(.+?)([^0-9\s]+(?:ビル|マンション|アパート|ハイツ|コーポ|館|棟|タワー|プラザ|センター|会館|ホール|ヴィラ|レジデンス|パレス|コート|テラス|ガーデン|ハウス).*)',
        # 階数表記（明確に○階、○F等）
        r'(.+?)(\s*\d+[階F].*)',
        # 号室表記（明確に○号室、○号等で終わる）
        r'(.+?)(\s*[0-9A-Za-z]+号室?\s*$)',
        # 括弧内の建物情報
        r'(.+?)(\s*[（(].+[)）]\s*$)',
    ]
    
    建物名分割済み = False
    for pattern in building_patterns:
        match = re.search(pattern, 残り住所)
        if match:
            住所1 = match.group(1).strip()
            住所2 = match.group(2).strip()
            建物名分割済み = True
            break
    
    # 建物名パターンに一致しない場合は、番地のみで分割しない
    if not 建物名分割済み:
        # **修正：番地情報は住所1に含める**
        # 以下のパターンは住所1にまとめる
        
        # 危険なパターン（分割しない）：
        # - 単純な番地（1-2-3、1丁目2番3号など）
        # - 番地のみの組み合わせ
        
        # より厳格な建物名判定
        strict_building_patterns = [
            # アルファベット+数字の組み合わせで明確に部屋番号と分かるもの
            r'(.+?)(\s*[A-Za-z]\d+\s*$)',  # A101、B205等
            # 区画・ブロック名（カタカナ）+ 番号
            r'(.+?)(\s*[ア-ヴ]+[0-9]+\s*$)',  # アルファ101等
        ]
        
        for pattern in strict_building_patterns:
            match = re.search(pattern, 残り住所)
            if match:
                住所1 = match.group(1).strip()
                住所2 = match.group(2).strip()
                建物名分割済み = True
                break
    
    # まだ分割されていない場合は全体を住所1に
    if not 建物名分割済み:
        住所1 = 残り住所.strip()
        住所2 = ""
    
    # **追加の安全チェック**：住所2が番地情報のみの場合は住所1に統合
    if 住所2:
        # 危険パターン：数字のみ、番地のみ、号のみなど
        危険パターン = [
            r'^\d+$',                    # 数字のみ
            r'^\d+号$',                  # ○号のみ
            r'^\d+-\d+$',               # ○-○のみ
            r'^\d+-\d+-\d+$',           # ○-○-○のみ
            r'^\d+番地?$',              # ○番地のみ
            r'^\d+丁目$',               # ○丁目のみ
            r'^\d+番\d+号?$',           # ○番○号のみ
        ]
        
        for pattern in 危険パターン:
            if re.match(pattern, 住所2.strip()):
                # 住所2を住所1に統合
                住所1 = f"{住所1} {住所2}".strip()
                住所2 = ""
                break
    
    return 郵便番号, 住所1, 住所2


def write_estimate_to_excel(data_or_template=None, output_filename=None):
    """見積書をExcelに出力する（係数対応版）"""
    
    # 引数の解析
    if data_or_template is None:
        import streamlit as st
        郵便番号, 住所1, 住所2 = get_customer_address_from_session()
        
        見積データ = {
            "見積No": st.session_state.get("見積No", ""),
            "案件名": st.session_state.get("案件名", ""),
            "発行日": st.session_state.get("発行日", datetime.today()),
            "顧客会社名": st.session_state.get("選択された顧客会社名", ""),
            "顧客部署名": st.session_state.get("選択された顧客部署名", ""),
            "顧客担当者": st.session_state.get("選択された顧客担当者", ""),
            "郵便番号": 郵便番号,
            "住所1": 住所1,
            "住所2": 住所2,
            "発行者名": st.session_state.get("発行者名", ""),
            "備考": st.session_state.get("備考", ""),
            "明細リスト": st.session_state.get("明細リスト", []),
            "係数機能使用": st.session_state.get("係数機能使用", False)
        }
        保存先ファイル名 = f"{見積データ['見積No']}見積書_{見積データ.get('案件名', '案件名未設定')}.xlsx"
    elif isinstance(data_or_template, dict):
        見積データ = data_or_template
        見積No = 見積データ.get('見積No', 'temp')
        案件名 = 見積データ.get('案件名', '案件名未設定')
        保存先ファイル名 = output_filename or f"{見積No}見積書_{案件名}.xlsx"
    else:
        import streamlit as st
        郵便番号, 住所1, 住所2 = get_customer_address_from_session()
        
        見積データ = {
            "見積No": st.session_state.get("見積No", ""),
            "案件名": st.session_state.get("案件名", ""),
            "発行日": st.session_state.get("発行日", datetime.today()),
            "顧客会社名": st.session_state.get("選択された顧客会社名", ""),
            "顧客部署名": st.session_state.get("選択された顧客部署名", ""),
            "顧客担当者": st.session_state.get("選択された顧客担当者", ""),
            "郵便番号": 郵便番号,
            "住所1": 住所1,
            "住所2": 住所2,
            "発行者名": st.session_state.get("発行者名", ""),
            "備考": st.session_state.get("備考", ""),
            "明細リスト": st.session_state.get("明細リスト", []),
            "係数機能使用": st.session_state.get("係数機能使用", False)
        }
        保存先ファイル名 = data_or_template
    
    # 係数機能使用状況に応じてテンプレートを選択
    係数機能使用 = 見積データ.get("係数機能使用", False)
    
    if 係数機能使用:
        template_path = "estimate_templat_keisuu.xlsx"  # 係数対応テンプレート
    else:
        template_path = "estimate_template.xlsx"  # 通常テンプレート
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"テンプレートファイルが見つかりません: {template_path}")
    
    wb = load_workbook(template_path)
    ws = wb.active

    def safe_write(ws, row, col, value):
        """結合セルに対応した書き込み"""
        for merged in ws.merged_cells.ranges:
            if (merged.min_row <= row <= merged.max_row and 
                merged.min_col <= col <= merged.max_col):
                ws.cell(merged.min_row, merged.min_col, value)
                return
        ws.cell(row, col, value)

    # 見積データから各項目を取得
    見積No = 見積データ.get("見積No", "")
    案件名 = 見積データ.get("案件名", "")
    顧客会社名 = 見積データ.get("顧客会社名", "")
    顧客部署名 = 見積データ.get("顧客部署名", "")
    顧客担当者 = 見積データ.get("顧客担当者", "")
    
    # 住所情報を取得
    郵便番号 = 見積データ.get("郵便番号", "")
    住所1 = 見積データ.get("住所1", "")
    住所2 = 見積データ.get("住所2", "")
    
    if not (郵便番号 or 住所1 or 住所2):
        旧住所 = 見積データ.get("顧客住所", "")
        if 旧住所:
            郵便番号, 住所1, 住所2 = parse_address(旧住所)
    
    発行日 = 見積データ.get("発行日", datetime.today())
    発行者名 = 見積データ.get("発行者名", "")
    備考 = 見積データ.get("備考", "")
    明細リスト = 見積データ.get("明細リスト", [])

    # 案件情報の書き込み
    safe_write(ws, 6, 2, 顧客会社名)
    
    # 郵便番号に〒マークを追加
    if 郵便番号:
        郵便番号表示 = f"〒{郵便番号}" if not 郵便番号.startswith("〒") else 郵便番号
    else:
        郵便番号表示 = ""
    
    safe_write(ws, 8, 2, 郵便番号表示)
    safe_write(ws, 8, 4, 住所1)
    safe_write(ws, 9, 4, 住所2)
    safe_write(ws,10, 4, 顧客担当者)
    # 発行日、見積No、発行者名の書き込み（係数対応）
    if 係数機能使用:
        # 係数対応テンプレートの場合（N列）
        safe_write(ws, 1, 14, 発行日)    # N1: 発行日
        safe_write(ws, 2, 14, 見積No)    # N2: 見積No
        safe_write(ws, 13, 14, 発行者名)  # N13: 発行者名
    else:
        # 通常テンプレートの場合（M列）
        safe_write(ws, 1, 13, 発行日)    # M1: 発行日
        safe_write(ws, 2, 13, 見積No)    # M2: 見積No
        safe_write(ws, 13, 13, 発行者名)  # M13: 発行者名
    safe_write(ws,15, 3, 案件名)
    safe_write(ws,43, 2, 備考)

    # 明細データの書き込み（係数対応版）
    明細開始行 = 18
    表示行カウンタ = 0
    
    for i, item in enumerate(明細リスト):
        current_row = 明細開始行 + 表示行カウンタ
        
        # 分類項目かどうかを判定
        is_category = item.get("分類", False)
        
        if is_category:
            # 分類項目の場合：番号なし、品名のみ太字で表示
            safe_write(ws, current_row, 2, "")                    # B列: 番号なし
            safe_write(ws, current_row, 3, item["品名"])          # C列: 分類名
            
            if 係数機能使用:
                # 係数対応テンプレートの場合
                safe_write(ws, current_row, 8, "")               # H列: 数量なし
                safe_write(ws, current_row, 9, "")               # I列: 単位なし
                safe_write(ws, current_row, 10, "")              # J列: 係数なし
                safe_write(ws, current_row, 11, "")              # K列: 単価なし
                safe_write(ws, current_row, 13, "")              # M列: 金額なし
                safe_write(ws, current_row, 15, "")              # O列: 備考なし
            else:
                # 通常テンプレートの場合
                safe_write(ws, current_row, 8, "")               # H列: 数量なし
                safe_write(ws, current_row, 9, "")               # I列: 単位なし
                safe_write(ws, current_row, 10, "")              # J列: 単価なし
                safe_write(ws, current_row, 12, "")              # L列: 金額なし
                safe_write(ws, current_row, 14, "")              # N列: 備考なし
            
            # 分類名を太字にする（オプション）
            try:
                from openpyxl.styles import Font
                cell = ws.cell(current_row, 3)
                cell.font = Font(bold=True)
            except:
                pass
        else:
            # 商品項目の場合：商品番号を使用
            商品番号 = item.get("商品番号")
            if 商品番号 is not None:
                safe_write(ws, current_row, 2, 商品番号)         # B列: 商品番号
            else:
                safe_write(ws, current_row, 2, "")               # B列: 番号なし
            
            safe_write(ws, current_row, 3, item["品名"])          # C列: 品名
            safe_write(ws, current_row, 8, item["数量"])          # H列: 数量
            safe_write(ws, current_row, 9, item["単位"])          # I列: 単位
            
            if 係数機能使用:
                # 係数対応テンプレートの場合
                safe_write(ws, current_row, 10, item.get("係数", 1))      # J列: 係数
                safe_write(ws, current_row, 11, item["単価"])             # K列: 単価
                safe_write(ws, current_row, 15, item.get("備考", ""))     # O列: 備考
                
                # M列（金額）の数式設定：H*J*K（数量*係数*単価）
                if isinstance(item.get("金額"), (int, float)) and item.get("金額") != 0:
                    m_cell = ws.cell(current_row, 13)
                    m_cell.value = f"=H{current_row}*J{current_row}*K{current_row}"
            else:
                # 通常テンプレートの場合
                safe_write(ws, current_row, 10, item["単価"])             # J列: 単価
                safe_write(ws, current_row, 14, item.get("備考", ""))     # N列: 備考
                
                # L列（金額）の数式設定：H*J（数量*単価）
                if isinstance(item.get("金額"), (int, float)) and item.get("金額") != 0:
                    l_cell = ws.cell(current_row, 12)
                    l_cell.value = f"=H{current_row}*J{current_row}"
        
        表示行カウンタ += 1
    
    # 使用していない明細行をクリア
    for row in range(明細開始行 + 表示行カウンタ, 40):
        if 係数機能使用:
            # 係数対応テンプレートの場合
            for col in [2, 3, 8, 9, 10, 11, 13, 15]:  # B,C,H,I,J,K,M,O列をクリア
                safe_write(ws, row, col, "")
        else:
            # 通常テンプレートの場合
            for col in [2, 3, 8, 9, 10, 12, 14]:  # B,C,H,I,J,L,N列をクリア
                safe_write(ws, row, col, "")

    # 合計欄の数式設定（係数対応）
    if 係数機能使用:
        # 係数対応テンプレートの場合（M列が金額）
        ws.cell(40, 14).value = "=SUM(M18:M39)"      # N40: 小計
        ws.cell(41, 14).value = "=ROUNDDOWN(N40*0.1,0)"  # N41: 消費税
        ws.cell(42, 14).value = "=N40+N41"           # N42: 合計
    else:
        # 通常テンプレートの場合（L列が金額）
        ws.cell(40, 13).value = "=SUM(L18:L39)"      # M40: 小計
        ws.cell(41, 13).value = "=ROUNDDOWN(M40*0.1,0)"  # M41: 消費税
        ws.cell(42, 13).value = "=M40+M41"           # M42: 合計

    # ファイルを保存
    wb.save(保存先ファイル名)
    return True

# 旧バージョンとの互換性のための関数
def write_estimate_to_excel_legacy(
    見積No, 案件名, 顧客会社名, 顧客部署名, 顧客担当者, 顧客住所,
    発行日, 発行者名, 備考, 明細リスト, 保存先ファイル名
):
    """旧バージョンとの互換性のための関数"""
    郵便番号, 住所1, 住所2 = parse_address(顧客住所)
    
    見積データ = {
        "見積No": 見積No,
        "案件名": 案件名,
        "顧客会社名": 顧客会社名,
        "顧客部署名": 顧客部署名,
        "顧客担当者": 顧客担当者,
        "郵便番号": 郵便番号,
        "住所1": 住所1,
        "住所2": 住所2,
        "発行日": 発行日,
        "発行者名": 発行者名,
        "備考": 備考,
        "明細リスト": 明細リスト
    }
    return write_estimate_to_excel(見積データ, 保存先ファイル名)